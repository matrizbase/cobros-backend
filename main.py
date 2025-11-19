from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl import load_workbook
import os

app = FastAPI()

# Seguridad básica (CORS)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "Plantilla_Basedatos.xlsx"
TARGET_SHEET = "Base tel"   # <<< Nombre exacto de la hoja


@app.get("/")
def root():
    return {"status": "Servidor funcionando", "archivo": EXCEL_FILE}


@app.post("/subir_base/")
async def subir_base(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un .xlsx")

    with open(EXCEL_FILE, "wb") as f:
        f.write(await file.read())

    return {"mensaje": "Archivo actualizado exitosamente"}


@app.get("/buscar_numero/")
def buscar_numero(numero: str):
    """
    Busca un número telefónico en la hoja 'Base tel'
    """

    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(status_code=404, detail="No existe la base de datos")

    wb = load_workbook(EXCEL_FILE)
    if TARGET_SHEET not in wb.sheetnames:
        raise HTTPException(status_code=404, detail=f"No existe la hoja '{TARGET_SHEET}'")

    ws = wb[TARGET_SHEET]

    # Buscar número en toda la hoja
    for row in ws.iter_rows(values_only=True):
        if numero in [str(cell) for cell in row if cell is not None]:
            return {"encontrado": True, "fila": list(row)}

    return {"encontrado": False, "fila": None}


@app.get("/obtener_todos/")
def obtener_todos():
    """
    Devuelve todos los datos de 'Base tel'
    """

    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(status_code=404, detail="No existe el archivo")

    wb = load_workbook(EXCEL_FILE)
    ws = wb[TARGET_SHEET]

    data = [list(row) for row in ws.iter_rows(values_only=True)]

    return {"datos": data}

